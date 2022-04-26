import pyautogui as py
import openpyxl as xl
import shutil
import time
import smtplib
from email.message import EmailMessage

def localizationPositionMouse():
    py.alert("Você entrou no modo de Automação. "
                    "Te daremos 5 seconds para colocar o mouse na posição desejada!!")
    py.sleep(5)
    print(py.position())

def enviarEmailTexto(email, emailReceptor, senha, texto):
    msg = EmailMessage()
    msg['Subject'] = 'Email Automatizado...'
    msg['From'] = email
    msg['To'] = emailReceptor
    msg.set_content(texto)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(email, senha)
        smtp.send_message(msg)

# ---------------------------------------------------------- S.O
def toAreaDeTrabalho():
    py.hotkey("winleft", "d")

def openWPSOfiice():
    toAreaDeTrabalho()
    py.press("winleft")
    time.sleep(2)
    py.write("WPS OFFICE")
    py.press("enter")
    time.sleep(10)

def openExplorerFiles():
    # Abrindo o Explorador de Arquivos
    py.press("winleft")
    time.sleep(1.5)
    py.write("Explorador de Arquivos")
    time.sleep(5)
    py.press("enter")

# ---------------------------------------------------------- Excel

def criarPlanilha(nomePlanilha):
    planilha = xl.Workbook()
    salvarPlanilha(planilha, nomePlanilha, False)
    print("Planilha Criada!")

def carregarPlanilha(caminho):
    planilha = xl.load_workbook(caminho)
    print('planilha Carregada')
    return planilha

def salvarPlanilha(planilha, nome, mostrarMsg):
    planilha.save(nome)
    if mostrarMsg:
        print("Planilha Salva!")

def mesclarCelulas(planilha, celulas, linhaInicial, linhaFinal, colunaInicial, colunaFinal):
    clls = planilha.active
    clls.merge_cells(celulas)
    clls.merge_cells(start_row=linhaInicial, start_column=colunaInicial, end_row=linhaFinal, end_column=colunaFinal)
    print('Células Mescladas')

def moverPlanilha(caminho, novoCaminho):
    shutil.move(caminho, novoCaminho)
    print("Planilha movida")

# ------------------------------------------------ git

class git():

    def addCommitAnPush(self):
        self.abrirProjectAndGit()
        self.addA()
        print("Add Sucess")
        self.abrirProjectAndGit()
        self.commit()
        print("Commit Sucess")
        self.abrirProjectAndGit()
        self.push()
        print("Push Sucess")

    def abrirProjectAndGit(self):
        py.moveTo(x=88, y=93)
        py.leftClick()
        py.rightClick()
        time.sleep(1)
        py.press("down", 20)
        time.sleep(1)
        py.press("right")

    def addA(self):
        time.sleep(1)
        py.press("down")
        time.sleep(1)
        py.press("enter")
        time.sleep(5)

    def commit(self):
        time.sleep(1)
        py.press("enter")
        time.sleep(1)
        py.write("commitMsg")
        time.sleep(1)
        py.moveTo(x=73, y=430)
        py.leftClick()
        time.sleep(5)
        py.moveTo(x=345, y=71)
        py.leftClick()
        py.moveTo(x=10, y=83)
        py.leftClick()
        time.sleep(5)

    def push(self):
        py.moveTo(x=608, y=293)
        py.leftClick()
        time.sleep(1)
        py.moveTo(x=717, y=569)
        py.leftClick()
