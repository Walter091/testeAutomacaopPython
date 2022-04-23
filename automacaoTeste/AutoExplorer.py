import pyautogui as py
import openpyxl
from pyautogui import press, leftClick, rightClick, doubleClick, moveTo
from nucleo import InfoMsg, Atalhos_Teclado
import AutoBase
from time import sleep
from pyperclip import paste

class TesteAutomacao():

    NumFile = None
    NameFile = "None"
    Status = "Original"

    def Index(self):
        #mensagem alert
        InfoMsg.msgInicializacao("Automação iniciada. Por favor, não use o teclado "
                        "e nem mecha no mouse, para evitar erros!")

        #Abrindo a Aréa de Trabalho
        AutoBase.toAreaDeTrabalho()
        #Abrindo o Explorador de Arquivos
        AutoBase.openExplorerFiles()
        #Maximizando a tela
        sleep(1)
        moveTo(x=858, y=152)
        leftClick()

        #Acessando e informando o caminho do arquivo zip
        self.openPastaOriginal()

        entra = True
        indice = 1
        while entra:
            if indice <= 11:
                if self.abrirArquivo(indice):
                    sleep(2)
                    self.salvarArquivo(indice)
                    sleep(2)
                    self.closeFile()
                    Atalhos_Teclado.altTab()
                    self.createPlanilha(indice, "Relatorio de execucao")
                    sleep(2)
                    indice = indice + 1
                    entra = True
                else:
                    InfoMsg.msgInicializacao("Impossivel Abrir arquivos .docx"
                                             " ou arquivos que não começam com um Númeral!!"
                                             "Pressione 'Enter' para Continuar.")
                    press("esc")
                    continue
            else:
                entra = False

        InfoMsg("Automação encerrada. Obrigado e volte Sempre!")

    def openPastaOriginal(self):
        self.searchFiles("Downloads", "RPA-Artigo")
        sleep(2)
        moveTo(x=255, y=158)
        doubleClick()

    def ordenarPorPdf(self):
        moveTo(x=569, y=103)
        leftClick()
        press("down")
        press("enter")

    def ordenarPorNome(self):
        moveTo(x=262, y=99)
        leftClick()

    def abrirArquivo(self, indice):
        sleep(2)
        conteudo = None
        if indice == 1:
            self.ordenarPorPdf()
            self.ordenarPorNome()
            press("down", 1)
            press("up")
            moveTo(x=524, y=133)
            rightClick()
            press("down", 5)
            press("enter")
            py.hotkey('ctrl', 'c') #Copia conteúdo selecionado
            sleep(2)
            conteudo = paste()
            py.typewrite(conteudo)  # Digita o texto da variável conteúdo
            file = str(conteudo)
            if self.validationArquivo(file):
                press("enter")
                sleep(1)
                press("enter")
                return True
            else:
                return False
        else:
            press("down", indice-1)
            press("apps")
            press("down", 5)
            press("enter")
            py.hotkey('ctrl', 'c') #Copia conteúdo selecionado
            conteudo = paste()
            py.typewrite(conteudo)  # Digita o texto da variável conteúdo
            sleep(2)
            file = str(conteudo)
            if self.validationArquivo(file):
                press("enter")
                sleep(1)
                press("enter")
                return True
            else:
                return False


    def salvarArquivo(self, indice):
        sleep(5)
        moveTo(x=31, y=50)
        leftClick()
        sleep(3)
        moveTo(x=83, y=239)
        leftClick()
        sleep(4)
        py.write("Pagina " + NumFile + " - Modificado")
        sleep(3)
        press("enter")
        global Status
        Status = "documento alterado"

    def validationArquivo(self, file):
        #Verifica se è numerica e .pdf
        if file[0].isnumeric():
            if file.__contains__(".pdf"):
                self.setInfoNameFile(file)
                return True
        else:
            return False

    def setInfoNameFile(self, file):
        global NumFile
        global NameFIle
        if file[0:2].isnumeric():
            NumFile = file[0:2]
        else:
            NumFile = file[0]
        # Adiciona a variavel para posterior utilização
        NameFIle = file


    # ----------------------------------------------------------------------------

    def createPlanilha(self, indice, nomePlanilha):
        planilha = openpyxl.Workbook()
        planilha.create_sheet('Documentos')
        planilha.save(planilha)

        self.addInfoPlanilha(indice, planilha)
        print("Planilha Criada!")

        self.salvarPlanilha(nomePlanilha)

    def mesclarCelulas(self, planilha):
        #Mesclando as células do Nome do documento
        clls = planilha.active
        clls.merge_cells('A2:D2')
        clls.merge_cells(start_row=2, start_column=1, end_row=11, end_column=4)
        #Mesclando as células do Status do documento
        clls.merge_cells('E2:G2')
        clls.merge_cells(start_row=2, start_column=5, end_row=11, end_column=8)

    def addInfoPlanilha(self, indice, planilha):
        self.mesclarCelulas()
        planilha.load_workbook('Relatorio de execucao.xlsx')
        documento = planilha['Documentos']
        if indice == 1 :
            documento.append(["Nome do Documento", "Status"])

        documento.append([self.Status, self.NameFile])

    def salvarPlanilha(self, planilha, nome):
        planilha.save(nome + '.xlsx')
        print("Planilha Salva!")

    # def criarCabecalho(self):
    #     #Seleciona o campo do cabecalho
    #     sleep(5)
    #     self.mesclarCelulas()
    #     #Voltando pro começo da planilha
    #     moveTo(x=93, y=205)
    #     leftClick()
    #     #Selecionando
    #     with py.hold('shift'):
    #         press('right')
    #         press("down", 10)
    #
    #     #Clica em Inserir
    #     moveTo(x=323, y=46)
    #     leftClick()
    #     sleep(1)
    #     #Clica na setinha de lado
    #     moveTo(x=1016, y=100)
    #     leftClick()
    #     sleep(2)
    #     #Clica em cabecalho e rodapé
    #     moveTo(x=359, y=94)
    #     leftClick()
    #     sleep(2)
    #     #Clica em cabeçalho personalizado
    #     moveTo(x=662, y=277)
    #     leftClick()
    #     sleep(2)
    #     #Adiciona as colunas
    #     py.write("Nome do Documento")
    #     press("tab")
    #     py.write("Status")
    #     sleep(2)
    #     moveTo(x=695, y=525)
    #     leftClick()
    #     sleep(2)
    #     #confirma a criação
    #     moveTo(x=622, y=550)
    #     leftClick()
    #     sleep(2)

    # ----------------------------------------------------------------------------

    def searchFiles(self, nameDirectory, nameFile):
        #Acessando o diretorio
        moveTo(x=615, y=68)
        leftClick()
        sleep(2)
        py.write(nameDirectory)
        press("enter")
        #Pesquisando e acessando o arquivo
        sleep(2)
        moveTo(x=724, y=64)
        leftClick()
        sleep(2)
        py.write(nameFile)
        press("enter")

    def closeFile(self):
        moveTo(x=345, y=11)
        sleep(1)
        leftClick()
        sleep(2)
